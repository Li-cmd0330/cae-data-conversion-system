"""
Excel导出模块 - 基于模板文件
严格按照模板格式填充key文件的数据
"""
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from typing import Dict, Any, List
import os
import shutil

# 导入可视化和验证模块
try:
    from core.validators.material_validator import DataValidator
    VALIDATION_AVAILABLE = True
except ImportError:
    VALIDATION_AVAILABLE = False

VISUALIZATION_AVAILABLE = False


class ExcelExporter:
    """
    Excel导出器 - 基于模板文件
    """
    
    def __init__(self, include_charts=True, include_validation=True):
        self.workbook = None
        self.include_charts = include_charts and VISUALIZATION_AVAILABLE
        self.include_validation = include_validation and VALIDATION_AVAILABLE
        
        # 模板文件路径：优先使用项目根目录下的模板文件
        script_dir = os.path.dirname(os.path.abspath(__file__))
        project_root = os.path.abspath(os.path.join(script_dir, "..", "..", "..", ".."))
        root_template = os.path.join(project_root, "材料数据模板-空.xlsx")
        local_template = os.path.join(script_dir, "材料数据模板-空.xlsx")
        self.TEMPLATE_PATH = root_template if os.path.exists(root_template) else local_template
        
        # 用于存储模板样式的字典
        self.template_styles = {}
        
        if self.include_charts:
            self.visualizer = DataVisualizer()
        if self.include_validation:
            self.validator = DataValidator()
    
    def _copy_cell_style(self, source_cell, target_cell):
        """
        复制单元格样式（从源单元格到目标单元格）
        """
        if source_cell.has_style:
            # 复制字体
            if source_cell.font:
                target_cell.font = Font(
                    name=source_cell.font.name,
                    size=source_cell.font.size,
                    bold=source_cell.font.bold,
                    italic=source_cell.font.italic,
                    vertAlign=source_cell.font.vertAlign,
                    underline=source_cell.font.underline,
                    strike=source_cell.font.strike,
                    color=source_cell.font.color
                )
            
            # 复制对齐
            if source_cell.alignment:
                target_cell.alignment = Alignment(
                    horizontal=source_cell.alignment.horizontal,
                    vertical=source_cell.alignment.vertical,
                    text_rotation=source_cell.alignment.text_rotation,
                    wrap_text=source_cell.alignment.wrap_text,
                    shrink_to_fit=source_cell.alignment.shrink_to_fit,
                    indent=source_cell.alignment.indent
                )
            
            # 复制填充
            if source_cell.fill:
                target_cell.fill = PatternFill(
                    fill_type=source_cell.fill.fill_type,
                    start_color=source_cell.fill.start_color,
                    end_color=source_cell.fill.end_color
                )
            
            # 复制边框
            if source_cell.border:
                target_cell.border = Border(
                    left=Side(style=source_cell.border.left.style, color=source_cell.border.left.color) if source_cell.border.left else None,
                    right=Side(style=source_cell.border.right.style, color=source_cell.border.right.color) if source_cell.border.right else None,
                    top=Side(style=source_cell.border.top.style, color=source_cell.border.top.color) if source_cell.border.top else None,
                    bottom=Side(style=source_cell.border.bottom.style, color=source_cell.border.bottom.color) if source_cell.border.bottom else None
                )
            
            # 复制数字格式
            if source_cell.number_format:
                target_cell.number_format = source_cell.number_format
    
    def _save_template_style(self, ws, row, col, key):
        """
        保存模板样式到字典中
        """
        cell = ws.cell(row=row, column=col)
        self.template_styles[key] = cell
    
    def _apply_template_style(self, ws, row, col, style_key):
        """
        应用保存的模板样式
        """
        if style_key in self.template_styles:
            source_cell = self.template_styles[style_key]
            target_cell = ws.cell(row=row, column=col)
            self._copy_cell_style(source_cell, target_cell)
    
    def _set_cell_with_style(self, ws, row, col, value, style_row=None, style_col=None):
        """
        设置单元格值并应用样式
        如果提供了style_row和style_col，则从该位置复制样式
        """
        target_cell = ws.cell(row=row, column=col)
        target_cell.value = value
        
        # 如果提供了样式源位置，则复制样式
        if style_row is not None and style_col is not None:
            style_cell = ws.cell(row=style_row, column=style_col)
            self._copy_cell_style(style_cell, target_cell)
    
    def create_material_template(self, material_data: Dict[str, Any], output_path: str):
        """
        创建材料数据Excel文件（基于模板）
        """
        # 检查模板文件是否存在
        if not os.path.exists(self.TEMPLATE_PATH):
            print(f"警告: 模板文件 {self.TEMPLATE_PATH} 不存在，使用默认格式")
            self._create_default_format(material_data, output_path)
            return
        
        # 复制模板文件
        shutil.copy2(self.TEMPLATE_PATH, output_path)
        
        # 打开复制的文件并填充数据
        self.workbook = openpyxl.load_workbook(output_path)
        
        # 填充各个工作表的数据
        self._fill_material_info(material_data)
        self._fill_mechanical_properties(material_data)
        self._fill_constitutive_data(material_data)
        self._fill_thermal_properties(material_data)
        self._fill_damage_data(material_data)
        
        # 添加数据验证报告和可视化（作为新工作表）
        if self.include_validation:
            self._add_validation_sheet(material_data)
        
        if self.include_charts and 'FSTRES' in material_data:
            self._add_visualization_sheet(material_data)
        
        # 保存文件
        self.workbook.save(output_path)
    
    def _fill_material_info(self, material_data: Dict[str, Any]):
        """填充材料基本信息工作表"""
        if "材料基本信息" not in self.workbook.sheetnames:
            return
        
        ws = self.workbook["材料基本信息"]
        
        # 根据模板结构：
        # 第1行：材料一级分类 | ... | 元素名称 | 下限 | 上限 | 实测值1 | 实测值2
        # 第2行：材料二级分类 | ... | Cu
        # 第3行：材料名称 | ... | Mg
        # 第1列包含标签（材料一级分类、材料二级分类、材料名称等）
        # 第3列包含元素名称（Cu, Mg, Mn, Fe, Si, Ni, Zn, Ti, Al）
        
        # 保存模板第2行第2列的样式作为参考
        self._save_template_style(ws, 2, 2, 'material_info_value')
        
        # 清除第2-11行第2列和第4列及以后的数据（保留第1行表头、第1列标签和第3列元素名称）
        for row_idx in range(2, 12):
            # 清除第2列
            ws.cell(row=row_idx, column=2).value = None
            # 清除第4列及以后（保留第3列的元素名称）
            for col in range(4, ws.max_column + 1):
                ws.cell(row=row_idx, column=col).value = None
        
        # 清除第12行及以后的所有数据
        for row_idx in range(12, 101):
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col)
                cell.value = None
        
        # 材料名称填在第3行第2列（应用样式）
        if 'MTNAME' in material_data:
            material_name = material_data['MTNAME'].get('name', '')
            self._set_cell_with_style(ws, 3, 2, material_name, 2, 2)
        
        # 单位系统可以填在合适的位置（如果模板中有预置的行）
        # 检查是否有"单位系统"标签
        unit_row = None
        for row_idx in range(1, 12):
            if ws.cell(row=row_idx, column=1).value and "单位" in str(ws.cell(row=row_idx, column=1).value):
                unit_row = row_idx
                break
        
        if 'UNIT' in material_data:
            unit_value = material_data['UNIT']
            if unit_row:
                self._set_cell_with_style(ws, unit_row, 2, unit_value, 2, 2)
            else:
                # 如果没有预置的单位系统行，添加到第12行
                self._set_cell_with_style(ws, 12, 1, "单位系统", 2, 1)
                self._set_cell_with_style(ws, 12, 2, unit_value, 2, 2)
    
    def _fill_mechanical_properties(self, material_data: Dict[str, Any]):
        """填充力学性能工作表"""
        if "力学性能" not in self.workbook.sheetnames:
            return
        
        ws = self.workbook["力学性能"]
        
        # 模板第2行已经预置了属性名称，我们需要保留第2行的属性名称列(1,4,7,10,12,14,17)
        # 只清除第2行的值列和温度列，以及第3行及以后的所有数据
        
        # 保存模板样式（包括属性名称列）
        for col in [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18]:
            self._save_template_style(ws, 2, col, f'mech_style_{col}')
        
        # 统一温度列样式：使用第6列（泊松比的温度列）作为所有温度列的样式参考
        # 这样可以确保所有温度列（第3、6、9、16列）样式一致
        self._save_template_style(ws, 2, 6, 'mech_temp_unified')
        
        # 保存第2行的属性名称以便后续填充
        property_names = {
            1: ws.cell(row=2, column=1).value,   # 杨氏模量
            4: ws.cell(row=2, column=4).value,   # 泊松比
            7: ws.cell(row=2, column=7).value,   # 弹性极限
            10: ws.cell(row=2, column=10).value, # 极限拉伸强度
            12: ws.cell(row=2, column=12).value, # 体积力
            14: ws.cell(row=2, column=14).value, # 硬度
            17: ws.cell(row=2, column=17).value  # 硬化类型
        }
        
        # 清除第3行及以后的所有数据
        for row in range(3, 101):
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                cell.value = None
        
        # 清除第2行的值列和温度列（保留属性名称列）
        # 根据模板结构：列1,4,7,10,12,14,17是属性名称，其他是值和温度
        for col in [2, 3, 5, 6, 8, 9, 11, 13, 15, 16, 18]:
            ws.cell(row=2, column=col).value = None
        
        # 现在在第2行开始填入key文件的真实数据
        row = 2
        
        # 杨氏模量 - 第1列（名称），第2列（值），第3列（温度）
        # 使用统一的温度列样式
        if 'YOUNG' in material_data:
            young_data = material_data['YOUNG']
            curve_data = young_data.get('curve_data', [])
            
            if curve_data:
                # 有曲线数据，填充多行
                for idx, point in enumerate(curve_data):
                    self._set_cell_with_style(ws, row + idx, 1, property_names[1], 2, 1)
                    self._set_cell_with_style(ws, row + idx, 2, point.get('value', 0), 2, 2)
                    # 使用统一的温度列样式
                    self._set_cell_with_style(ws, row + idx, 3, point.get('temperature', 20), 2, 6)
            else:
                # 常数值
                young_value = young_data.get('value', 0)
                self._set_cell_with_style(ws, row, 1, property_names[1], 2, 1)
                self._set_cell_with_style(ws, row, 2, young_value, 2, 2)
                # 使用统一的温度列样式
                self._set_cell_with_style(ws, row, 3, 20, 2, 6)  # 默认温度20℃
        
        # 泊松比 - 第4列（名称），第5列（值），第6列（温度）
        if 'POISON' in material_data:
            poison_data = material_data['POISON']
            curve_data = poison_data.get('curve_data', [])
            
            if curve_data:
                # 有曲线数据，填充多行
                for idx, point in enumerate(curve_data):
                    self._set_cell_with_style(ws, row + idx, 4, property_names[4], 2, 4)
                    self._set_cell_with_style(ws, row + idx, 5, point.get('value', 0), 2, 5)
                    self._set_cell_with_style(ws, row + idx, 6, point.get('temperature', 20), 2, 6)
            else:
                # 常数值
                poison_value = poison_data.get('value', 0)
                self._set_cell_with_style(ws, row, 4, property_names[4], 2, 4)
                self._set_cell_with_style(ws, row, 5, poison_value, 2, 5)
                self._set_cell_with_style(ws, row, 6, 20, 2, 6)  # 默认温度20℃
        
        # 弹性极限 - 第7列（名称），第8列（值），第9列（温度）
        if 'ELRST' in material_data:
            elrst_value = material_data['ELRST'].get('value', 0)
            self._set_cell_with_style(ws, row, 7, property_names[7], 2, 7)
            self._set_cell_with_style(ws, row, 8, elrst_value, 2, 8)
            # 使用统一的温度列样式
            self._set_cell_with_style(ws, row, 9, 20, 2, 6)  # 默认温度20℃
        
        # 极限拉伸强度 - 第10列（名称），第11列（值）
        if 'UTSDAT' in material_data:
            utsdat_value = material_data['UTSDAT'].get('value', 0)
            self._set_cell_with_style(ws, row, 10, property_names[10], 2, 10)
            self._set_cell_with_style(ws, row, 11, utsdat_value, 2, 11)
        
        # 体积力 - 第12列（名称），第13列（值）
        if 'FPERV' in material_data:
            fperv_value = material_data['FPERV'].get('value', 0)
            self._set_cell_with_style(ws, row, 12, property_names[12], 2, 12)
            self._set_cell_with_style(ws, row, 13, fperv_value, 2, 13)
        
        # 硬度 - 第14列（名称），第15列（值），第16列（温度）
        if 'HDNPHA' in material_data:
            hdnpha_value = material_data['HDNPHA'].get('value', 0)
            self._set_cell_with_style(ws, row, 14, property_names[14], 2, 14)
            self._set_cell_with_style(ws, row, 15, hdnpha_value, 2, 15)
            # 使用统一的温度列样式
            self._set_cell_with_style(ws, row, 16, 20, 2, 6)  # 默认温度20℃
        
        # 硬化类型 - 第17列（名称），第18列（值）
        if 'HDNRUL' in material_data:
            hdnrul_val = material_data['HDNRUL'].get('type', 0)
            hardening_type = "各向同性硬化" if hdnrul_val == 0 else f"类型{hdnrul_val}"
            self._set_cell_with_style(ws, row, 17, property_names[17], 2, 17)
            self._set_cell_with_style(ws, row, 18, hardening_type, 2, 18)
    
    def _fill_constitutive_data(self, material_data: Dict[str, Any]):
        """
        填充本构工作表（FSTRES数据）
        
        FSTRES格式说明：
        1行：FSTRES + 类型
        2行：应变点数k + 应变速率组数m + 温度组数n
        3行：应变值（k个点）
        4行：应变速率值（m组）
        5行：温度值（n组）
        6行起：应力值，共n*m*k个
        顺序：((STRESS(N,M,K), K=1,k), M=1,m), N=1,n)
        即：外层循环温度N，中层循环应变速率M，内层循环应变K
        """
        # 模板中的工作表名称是"塑性"
        if "塑性" not in self.workbook.sheetnames:
            return
        
        if 'FSTRES' not in material_data:
            return
        
        ws = self.workbook["塑性"]
        
        # 保存模板第2行的样式
        for col in range(1, 5):
            self._save_template_style(ws, 2, col, f'constitutive_style_{col}')
        
        # 彻底清除模板中第2行及以后的所有示例数据（保留第1行表头）
        # 清除到第1000行以确保所有模板数据都被清除（本构数据可能很多）
        for row_idx in range(2, 1001):
            for col in range(1, 5):  # 只清除前4列：应变率、温度、应变、应力
                cell = ws.cell(row=row_idx, column=col)
                cell.value = None
        
        fstres = material_data['FSTRES']
        strain_data = fstres.get('strain_data', [])
        strain_rate_data = fstres.get('strain_rate_data', [])
        temp_data = fstres.get('temperature_data', [])
        stress_data = fstres.get('stress_data', [])  # 三维数组：[温度][应变速率][应变]
        
        num_strain = fstres.get('num_strain', len(strain_data))
        num_rate = fstres.get('num_rate', len(strain_rate_data))
        num_temp = fstres.get('num_temp', len(temp_data))
        
        # 模板第1行是表头：应变率 | 温度(℃) | 应变 | 应力(MPa) | ...
        # 从第2行开始填充数据
        
        row = 2
        data_count = 0
        
        # 按照FSTRES格式填充：外层温度N，中层应变速率M，内层应变K
        # stress_data结构：[温度索引N][应变速率索引M][应变索引K]
        
        if isinstance(stress_data, list) and len(stress_data) > 0:
            # 外层循环：温度
            for temp_idx, temp in enumerate(temp_data):
                if temp_idx < len(stress_data):
                    temp_stress_data = stress_data[temp_idx]
                    
                    # 中层循环：应变速率
                    for rate_idx, strain_rate in enumerate(strain_rate_data):
                        if rate_idx < len(temp_stress_data):
                            rate_stress_data = temp_stress_data[rate_idx]
                            
                            # 内层循环：应变
                            for strain_idx, strain in enumerate(strain_data):
                                if strain_idx < len(rate_stress_data):
                                    # 填充：应变率 | 温度 | 应变 | 应力（应用样式）
                                    self._set_cell_with_style(ws, row, 1, strain_rate, 2, 1)
                                    self._set_cell_with_style(ws, row, 2, temp, 2, 2)
                                    self._set_cell_with_style(ws, row, 3, strain, 2, 3)
                                    self._set_cell_with_style(ws, row, 4, rate_stress_data[strain_idx], 2, 4)
                                    
                                    row += 1
                                    data_count += 1
    
    def _fill_thermal_properties(self, material_data: Dict[str, Any]):
        """填充热学性能工作表"""
        # 模板中的工作表名称是"热物性"
        if "热物性" not in self.workbook.sheetnames:
            return
        
        ws = self.workbook["热物性"]
        
        # 保存模板样式（包括属性名称列）- 必须在清除之前保存
        for col in range(1, 23):
            self._save_template_style(ws, 2, col, f'thermal_style_{col}')
        
        # 统一温度列样式：使用第6列（比热容的温度列）作为所有温度列的样式参考
        self._save_template_style(ws, 2, 6, 'thermal_temp_unified')
        
        # 保存第2行的属性名称以便后续填充
        property_names = {
            1: ws.cell(row=2, column=1).value,   # 热导率
            4: ws.cell(row=2, column=4).value,   # 比热容
            7: ws.cell(row=2, column=7).value,   # 热膨胀系数
            11: ws.cell(row=2, column=11).value, # 密度
            14: ws.cell(row=2, column=14).value, # 功热转换系数
            17: ws.cell(row=2, column=17).value, # 热辐射系数
            19: ws.cell(row=2, column=19).value  # 摩擦系数
        }
        
        # 获取工作表的真实最大列数（遍历所有行找到最大列）
        max_col = ws.max_column
        if max_col < 30:
            max_col = 30  # 确保清除足够多的列
        
        # 彻底清除第1行第22列及以后的所有多余列（表头和数据）
        for row_idx in range(1, 101):
            for col in range(22, max_col + 10):  # 多清除10列确保彻底
                cell = ws.cell(row=row_idx, column=col)
                cell.value = None
        
        # 清除第3行及以后第1-21列的数据
        for row_idx in range(3, 101):
            for col in range(1, 22):
                cell = ws.cell(row=row_idx, column=col)
                cell.value = None
        
        # 清除第2行的值列和温度列（保留属性名称列）
        # 根据模板结构：列1,4,7,11,14,17,19是属性名称，其他是值和温度
        # 特别说明：第21列（U列）用户要求必须为空，所以也清除
        for col in [2, 3, 5, 6, 8, 9, 10, 12, 13, 15, 16, 18, 20, 21]:
            ws.cell(row=2, column=col).value = None
        
        # 模板第1行：属性 | 值 | 温度（℃） | 属性 | 值 | 温度（℃） | ...
        # 填充第2行开始的真实数据
        row = 2
        
        # 热导率 - 第1列（名称），第2列（值），第3列（温度）
        # 使用统一的温度列样式
        if 'THRCND' in material_data:
            thrcnd_data = material_data['THRCND']
            curve_data = thrcnd_data.get('curve_data', [])
            
            if curve_data:
                # 有曲线数据，填充多行
                for idx, point in enumerate(curve_data):
                    self._set_cell_with_style(ws, row + idx, 1, property_names[1], 2, 1)
                    self._set_cell_with_style(ws, row + idx, 2, point.get('value', 0), 2, 2)
                    # 使用统一的温度列样式
                    self._set_cell_with_style(ws, row + idx, 3, point.get('temperature', 20), 2, 6)
            else:
                # 常数值
                thrcnd_value = thrcnd_data.get('value', 0)
                self._set_cell_with_style(ws, row, 1, property_names[1], 2, 1)
                self._set_cell_with_style(ws, row, 2, thrcnd_value, 2, 2)
                # 使用统一的温度列样式
                self._set_cell_with_style(ws, row, 3, 20, 2, 6)  # 默认温度20℃
        
        # 比热容 - 第4列（名称），第5列（值），第6列（温度）
        if 'HEATCP' in material_data:
            heatcp_data = material_data['HEATCP']
            curve_data = heatcp_data.get('curve_data', [])
            
            if curve_data:
                # 有曲线数据，填充多行
                for idx, point in enumerate(curve_data):
                    self._set_cell_with_style(ws, row + idx, 4, property_names[4], 2, 4)
                    self._set_cell_with_style(ws, row + idx, 5, point.get('value', 0), 2, 5)
                    self._set_cell_with_style(ws, row + idx, 6, point.get('temperature', 20), 2, 6)
            else:
                # 常数值
                heatcp_value = heatcp_data.get('value', 0)
                self._set_cell_with_style(ws, row, 4, property_names[4], 2, 4)
                self._set_cell_with_style(ws, row, 5, heatcp_value, 2, 5)
                self._set_cell_with_style(ws, row, 6, 20, 2, 6)  # 默认温度20℃
        
        # 热膨胀系数 - 第7列（名称），第8列（系数），第9列（温度），第10列（参考温度）
        if 'EXPAND' in material_data:
            expand_data = material_data['EXPAND']
            curve_data = expand_data.get('curve_data', [])
            ref_temp = expand_data.get('reference_temp', 20)
            
            if curve_data:
                # 有曲线数据，填充多行
                for idx, point in enumerate(curve_data):
                    self._set_cell_with_style(ws, row + idx, 7, property_names[7], 2, 7)
                    self._set_cell_with_style(ws, row + idx, 8, point.get('coefficient', 0), 2, 8)
                    # 使用统一的温度列样式
                    self._set_cell_with_style(ws, row + idx, 9, point.get('temperature', 20), 2, 6)
                    self._set_cell_with_style(ws, row + idx, 10, ref_temp, 2, 10)
            else:
                # 常数值
                coeff = expand_data.get('coefficient', 0)
                self._set_cell_with_style(ws, row, 7, property_names[7], 2, 7)
                self._set_cell_with_style(ws, row, 8, coeff, 2, 8)
                # 使用统一的温度列样式
                self._set_cell_with_style(ws, row, 9, 20, 2, 6)  # 默认温度20℃
                self._set_cell_with_style(ws, row, 10, ref_temp, 2, 10)
        
        # 密度 - 第11列（名称），第12列（值），第13列（温度）
        if 'MASDEN' in material_data:
            masden_data = material_data['MASDEN']
            curve_data = masden_data.get('curve_data', [])
            
            if curve_data:
                # 有曲线数据，填充多行
                for idx, point in enumerate(curve_data):
                    self._set_cell_with_style(ws, row + idx, 11, property_names[11], 2, 11)
                    self._set_cell_with_style(ws, row + idx, 12, point.get('value', 0), 2, 12)
                    # 使用统一的温度列样式
                    self._set_cell_with_style(ws, row + idx, 13, point.get('temperature', 20), 2, 6)
            else:
                # 常数值
                masden_value = masden_data.get('value', 0)
                self._set_cell_with_style(ws, row, 11, property_names[11], 2, 11)
                self._set_cell_with_style(ws, row, 12, masden_value, 2, 12)
                # 使用统一的温度列样式
                self._set_cell_with_style(ws, row, 13, 20, 2, 6)  # 默认温度20℃
        
        # 功热转换系数 - 第14列（名称），第15列（值），第16列（温度）
        if 'FRAE2H' in material_data:
            frae2h_value = material_data['FRAE2H'].get('value', 0)
            self._set_cell_with_style(ws, row, 14, property_names[14], 2, 14)
            self._set_cell_with_style(ws, row, 15, frae2h_value, 2, 15)
            # 使用统一的温度列样式
            self._set_cell_with_style(ws, row, 16, 20, 2, 6)  # 默认温度20℃
        
        # 热辐射系数 - 第17列（名称），第18列（值）
        if 'EMSVTY' in material_data:
            emsvty_data = material_data['EMSVTY']
            curve_data = emsvty_data.get('curve_data', [])
            
            if curve_data:
                # 有曲线数据，填充多行
                for idx, point in enumerate(curve_data):
                    self._set_cell_with_style(ws, row + idx, 17, property_names[17], 2, 17)
                    # 热辐射系数只有值，没有温度列
                    self._set_cell_with_style(ws, row + idx, 18, point.get('value', 0), 2, 18)
            else:
                # 常数值
                emsvty_value = emsvty_data.get('value', 0)
                self._set_cell_with_style(ws, row, 17, property_names[17], 2, 17)
                self._set_cell_with_style(ws, row, 18, emsvty_value, 2, 18)
        
        # 摩擦系数 - 第19列（名称），第20列（值）
        if 'FPERV' in material_data:
            fperv_value = material_data['FPERV'].get('value', 0)
            self._set_cell_with_style(ws, row, 19, property_names[19], 2, 19)
            self._set_cell_with_style(ws, row, 20, fperv_value, 2, 20)
        
        # 扩散系数 - 第21列（值）
        # 用户要求：U列（第21列）不应该有任何值，因此注释掉此部分
        # if 'DIFCOE' in material_data:
        #     difcoe_value = material_data['DIFCOE'].get('value', 0)
        #     self._set_cell_with_style(ws, row, 21, difcoe_value, 2, 21)
        
        # 删除工作表中第22列及以后的所有列（包括熔点等多余属性）
        # 先获取实际的最大列数，然后删除第22列及以后的所有列
        actual_max_col = ws.max_column
        if actual_max_col >= 22:
            # 删除从第22列开始的所有列
            cols_to_delete = actual_max_col - 21
            if cols_to_delete > 0:
                ws.delete_cols(22, cols_to_delete)
    
    def _fill_damage_data(self, material_data: Dict[str, Any]):
        """填充损伤工作表
        
        模板布局（按列组织）：
        第1行第1列：Cockcroft-Latham模型
        第2行第1列：断裂能量密度阈值(mJ/mm3)
        第2行第2列：值（damage_model=1的值填在这里）
        
        第1行第3列：归一化Cockcroft-Latham模型
        第2行第3列：断裂阈值系数
        第2行第4列：值（damage_model=0的值填在这里）
        
        第1行第7列：lode参数
        第1行第8列：温度
        第2行起：空白行用于手动填充数据
        """
        if "损伤" not in self.workbook.sheetnames:
            return
        
        ws = self.workbook["损伤"]
        
        # 保存模板样式（第2行的样式）
        for col in [1, 2, 3, 4]:
            self._save_template_style(ws, 2, col, f'damage_style_col_{col}')
        
        # 获取工作表的最大列数
        max_col = ws.max_column if ws.max_column > 10 else 10
        
        # 清除第2行的值列（第2、4列）
        ws.cell(row=2, column=2).value = None
        ws.cell(row=2, column=4).value = None
        
        # 清除lode参数和温度区域（第7-8列，所有行）
        for row_idx in range(1, 101):
            for col in [7, 8]:
                ws.cell(row=row_idx, column=col).value = None
        
        # 清除第3行及以后的其他列数据（第1-4列）
        for row_idx in range(3, 101):
            for col in range(1, 5):
                cell = ws.cell(row=row_idx, column=col)
                cell.value = None
        
        # 损伤模型 (FRCMOD)
        if 'FRCMOD' in material_data:
            damage_model = material_data['FRCMOD'].get('damage_model')
            
            # Cockcroft-Latham模型 (damage_model=1)
            if damage_model == 1:
                value = material_data['FRCMOD'].get('value')
                if value is not None:
                    self._set_cell_with_style(ws, 2, 2, value, 2, 2)
            
            # 归一化Cockcroft-Latham模型 (damage_model=0)
            elif damage_model == 0:
                value = material_data['FRCMOD'].get('value')
                if value is not None:
                    self._set_cell_with_style(ws, 2, 4, value, 2, 4)
        
        # lode参数和温度 - 创建两列简单表格
        # 第1行第7列：填充标题"lode参数"
        lode_header = ws.cell(row=1, column=7)
        lode_header.value = "lode参数"
        
        # 第1行第8列：填充标题"温度"
        temp_header = ws.cell(row=1, column=8)
        temp_header.value = "温度"
        
        # 设置标题样式（与第1行第1列相同）
        title_cell = ws.cell(row=1, column=1)
        if title_cell.has_style:
            for header_cell in [lode_header, temp_header]:
                header_cell.font = Font(
                    name=title_cell.font.name,
                    size=title_cell.font.size,
                    bold=title_cell.font.bold,
                    color=title_cell.font.color
                )
                header_cell.alignment = Alignment(
                    horizontal=title_cell.alignment.horizontal,
                    vertical=title_cell.alignment.vertical,
                    wrap_text=title_cell.alignment.wrap_text
                )
                header_cell.fill = PatternFill(
                    fill_type=title_cell.fill.fill_type,
                    start_color=title_cell.fill.start_color,
                    end_color=title_cell.fill.end_color
                )
        
        # 创建第2行的空白单元格（用于数据填充，直接复制模板样式包括颜色）
        # 第2行第7列（lode参数值）
        lode_data_cell = ws.cell(row=2, column=7)
        lode_data_cell.value = None  # 空白，用户可以手动填充
        # 完整复制第2行第2列的所有样式（包括背景颜色）
        self._copy_cell_style(ws.cell(2, 2), lode_data_cell)
        
        # 第2行第8列（温度值）
        temp_data_cell = ws.cell(row=2, column=8)
        temp_data_cell.value = None  # 空白，用户可以手动填充
        # 完整复制第2行第2列的所有样式（包括背景颜色）
        self._copy_cell_style(ws.cell(2, 2), temp_data_cell)
    
    def _create_default_format(self, material_data: Dict[str, Any], output_path: str):
        """
        如果模板不存在，使用默认格式创建
        """
        self.workbook = openpyxl.Workbook()
        
        if 'Sheet' in self.workbook.sheetnames:
            del self.workbook['Sheet']
        
        # 创建基本工作表
        ws = self.workbook.create_sheet("材料数据")
        
        # 设置表头样式
        header_font = Font(name='微软雅黑', size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        for col, header_text in [(1, "属性"), (2, "值"), (3, "英文名称")]:
            cell = ws.cell(row=1, column=col)
            cell.value = header_text
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # 设置数据行样式
        data_font = Font(name='微软雅黑', size=10)
        data_alignment = Alignment(horizontal='left', vertical='center')
        data_border = Border(
            left=Side(style='thin', color='DCDFE6'),
            right=Side(style='thin', color='DCDFE6'),
            top=Side(style='thin', color='DCDFE6'),
            bottom=Side(style='thin', color='DCDFE6')
        )
        
        row = 2
        if 'MTNAME' in material_data:
            for col, val in [(1, "材料名称"), (2, material_data['MTNAME'].get('name', ''))]:
                cell = ws.cell(row=row, column=col)
                cell.value = val
                cell.font = data_font
                cell.alignment = data_alignment
                cell.border = data_border
            row += 1
        
        if 'UNIT' in material_data:
            for col, val in [(1, "单位系统"), (2, material_data['UNIT'])]:
                cell = ws.cell(row=row, column=col)
                cell.value = val
                cell.font = data_font
                cell.alignment = data_alignment
                cell.border = data_border
            row += 1
        
        # 添加其他属性...
        
        # 设置列宽
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 30
        
        if self.include_validation:
            self._add_validation_sheet(material_data)
        
        if self.include_charts and 'FSTRES' in material_data:
            self._add_visualization_sheet(material_data)
        
        self.workbook.save(output_path)
    
    def _add_validation_sheet(self, material_data: Dict[str, Any]):
        """添加材料数据解析报告工作表"""
        try:
            validation_sheet = self.workbook.create_sheet("材料数据解析报告")
            
            row = 1
            cell = validation_sheet.cell(row=row, column=1)
            cell.value = "材料数据解析报告 | Material Data Analysis Report"
            cell.font = Font(name='微软雅黑', size=14, bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')
            validation_sheet.merge_cells(f'A{row}:C{row}')
            
            row += 2
            
            cell = validation_sheet.cell(row=row, column=1)
            cell.value = "数据统计信息"
            cell.font = Font(name='微软雅黑', size=12, bold=True)
            cell.fill = PatternFill(start_color="B4C7E7", end_color="B4C7E7", fill_type="solid")
            row += 1
            
            all_possible_props = [
                'MTNAME', 'UNIT', 'YOUNG', 'POISON', 'THRCND', 'HEATCP', 
                'MASDEN', 'EMSVTY', 'EXPAND', 'FRAE2H', 'FPERV', 'FSTRES',
                'HDNPHA', 'HDNRUL', 'FRCMOD', 'DIFCOE', 'ELRST', 'UTSDAT'
            ]
            
            present_props = [p for p in all_possible_props if p in material_data]
            completeness = (len(present_props) / len(all_possible_props)) * 100
            
            validation_sheet.cell(row=row, column=1).value = "数据完整度:"
            validation_sheet.cell(row=row, column=2).value = f"{completeness:.1f}%"
            validation_sheet.cell(row=row, column=2).font = Font(size=11, bold=True, color="00B050")
            row += 1
            
            validation_sheet.cell(row=row, column=1).value = "已解析属性数:"
            validation_sheet.cell(row=row, column=2).value = f"{len(present_props)} / {len(all_possible_props)}"
            row += 2
            
            prop_names = {
                'MTNAME': '材料名称 (Material Name)',
                'UNIT': '单位系统',
                'YOUNG': '杨氏模量 (Young\'s Modulus)',
                'POISON': '泊松比 (Poisson\'s Ratio)',
                'THRCND': '热导率 (Thermal Conductivity)',
                'HEATCP': '比热容 (Specific Heat)',
                'MASDEN': '密度 (Mass Density)',
                'EMSVTY': '热辐射系数 (Emissivity)',
                'EXPAND': '热膨胀系数 (Thermal Expansion Coefficient)',
                'FRAE2H': '功热转换系数 (Taylor-Quinney Coefficient)',
                'FPERV': '体积力 (Body Force)',
                'FSTRES': '本构模型 (Flow Stress Model)',
                'FRCMOD': '损伤模型 (Fracture Model)',
                'HDNPHA': '硬度 (Hardness)',
                'HDNRUL': '硬化类型 (Hardening Type)',
                'DIFCOE': '扩散系数',
                'ELRST': '弹性极限',
                'UTSDAT': '极限拉伸强度'
            }
            
            cell = validation_sheet.cell(row=row, column=1)
            cell.value = "已解析的属性列表"
            cell.font = Font(name='微软雅黑', size=11, bold=True)
            cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            row += 1
            
            for prop in present_props:
                prop_name = prop_names.get(prop, prop)
                validation_sheet.cell(row=row, column=1).value = f"[OK] {prop_name}"
                validation_sheet.cell(row=row, column=1).font = Font(color="00B050")
                
                if prop == 'MTNAME' and 'MTNAME' in material_data:
                    val = material_data['MTNAME'].get('name', '')
                    validation_sheet.cell(row=row, column=2).value = val
                elif prop == 'UNIT':
                    val = material_data.get('UNIT', '')
                    validation_sheet.cell(row=row, column=2).value = f"单位系统 {val}"
                elif prop == 'FSTRES' and 'FSTRES' in material_data:
                    fstres = material_data['FSTRES']
                    val = f"{fstres.get('num_strain', 0)}个应变点, {fstres.get('num_temp', 0)}个温度点"
                    validation_sheet.cell(row=row, column=2).value = val
                elif isinstance(material_data.get(prop), dict) and 'value' in material_data[prop]:
                    val = material_data[prop]['value']
                    validation_sheet.cell(row=row, column=2).value = f"{val:.6g}"
                
                row += 1
            
            validation_sheet.column_dimensions['A'].width = 50
            validation_sheet.column_dimensions['B'].width = 30
            validation_sheet.column_dimensions['C'].width = 20
            
        except Exception as e:
            print(f"添加解析报告失败: {e}")
    
    def _add_visualization_sheet(self, material_data: Dict[str, Any]):
        """添加数据可视化工作表"""
        try:
            chart_sheet = self.workbook.create_sheet("数据可视化")
            
            row = 1
            cell = chart_sheet.cell(row=row, column=1)
            cell.value = "数据可视化图表 | Data Visualization"
            cell.font = Font(name='微软雅黑', size=14, bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')
            chart_sheet.merge_cells(f'A{row}:F{row}')
            
            row += 2
            
            if 'FSTRES' in material_data:
                try:
                    fstres = material_data['FSTRES']
                    
                    cell = chart_sheet.cell(row=row, column=1)
                    cell.value = "流动应力曲线图 (应力-应变曲线)"
                    cell.font = Font(name='微软雅黑', size=11, bold=True)
                    row += 1
                    
                    cell = chart_sheet.cell(row=row, column=1)
                    desc = f"应变点数: {fstres.get('num_strain', 0)}, 温度点数: {fstres.get('num_temp', 0)}"
                    cell.value = desc
                    cell.font = Font(name='微软雅黑', size=9, color="666666")
                    row += 1
                    
                    chart_bytes = self.visualizer.create_flow_stress_chart(fstres)
                    
                    if chart_bytes:
                        from data_visualizer import DataVisualizer
                        DataVisualizer.image_to_excel(chart_bytes, chart_sheet, f'A{row}', 
                                                     width=600, height=400)
                        row += 25
                    else:
                        chart_sheet.cell(row=row, column=1).value = "图表生成失败：数据不足"
                        chart_sheet.cell(row=row, column=1).font = Font(color="FF0000")
                        row += 2
                        
                except Exception as e:
                    print(f"流动应力图生成失败: {e}")
                    chart_sheet.cell(row=row, column=1).value = f"图表生成失败: {str(e)}"
                    chart_sheet.cell(row=row, column=1).font = Font(color="FF0000")
                    row += 2
            
            try:
                cell = chart_sheet.cell(row=row, column=1)
                cell.value = "材料属性雷达图"
                cell.font = Font(name='微软雅黑', size=11, bold=True)
                row += 1
                
                radar_bytes = self.visualizer.create_statistics_chart(material_data)
                if radar_bytes:
                    from data_visualizer import DataVisualizer
                    DataVisualizer.image_to_excel(radar_bytes, chart_sheet, f'A{row}', 
                                                 width=500, height=500)
                else:
                    chart_sheet.cell(row=row, column=1).value = "图表生成失败：数据不足"
                    chart_sheet.cell(row=row, column=1).font = Font(color="FF0000")
                    
            except Exception as e:
                print(f"雷达图生成失败: {e}")
                chart_sheet.cell(row=row, column=1).value = f"雷达图生成失败: {str(e)}"
                chart_sheet.cell(row=row, column=1).font = Font(color="FF0000")
            
            for col in range(1, 7):
                chart_sheet.column_dimensions[get_column_letter(col)].width = 15
            
        except Exception as e:
            print(f"添加可视化图表失败: {e}")


if __name__ == "__main__":
    # 测试代码
    test_data = {
        'UNIT': 1,
        'MTNAME': {'id': 1, 'name': 'AL6061 Machining'},
        'FRAE2H': {'id': 1, 'value': 0.9},
        'FPERV': {'id': 1, 'value': 0.0},
        'YOUNG': {'id': 1, 'type': 0, 'value': 68900.0},
        'POISON': {'id': 1, 'type': 0, 'value': 0.3},
        'EXPAND': {'id': 1, 'type': 0, 'coefficient': 0.0, 'reference_temp': 20.0},
        'THRCND': {'id': 1, 'type': 0, 'value': 180.0},
        'HEATCP': {'id': 1, 'type': 0, 'value': 2.4},
        'EMSVTY': {'id': 1, 'type': 0, 'value': 0.25},
        'MASDEN': {'id': 1, 'type': 0, 'value': 2.7e-9},
        'HDNPHA': {'id': 1, 'type': 0, 'value': 0.0},
        'FSTRES': {
            'id': 1,
            'type': 2,
            'num_strain': 3,
            'num_temp': 2,
            'strain_rate': 3,
            'strain_data': [0.0, 0.05, 0.1],
            'temperature_data': [1000, 10000],
            'stress_data': [
                [20, 190, 345],
                [200, 380, 420]
            ]
        }
    }
    
    exporter = ExcelExporter(include_charts=False, include_validation=True)
    exporter.create_material_template(test_data, 'test_corrected.xlsx')
    print("\n基于模板的Excel文件已生成：test_corrected.xlsx")
